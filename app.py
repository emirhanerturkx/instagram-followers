import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import StaleElementReferenceException, TimeoutException
from urllib.parse import urlparse
import os
import time

workbook = openpyxl.load_workbook('excel_dosyasi.xlsx')
sheet = workbook.active

if sheet['A1'].value != 'Instagram Link':
    raise ValueError("Excel dosyasının başlık satırı 'Instagram Link' olmalıdır.")

followers = {}

chrome_service = ChromeService(executable_path='C:\\Users\\Emirhan Ertürk\\Desktop\\samet\\chromedriver.exe')
driver = webdriver.Chrome(service=chrome_service)

driver.get("https://www.instagram.com/")

username = "********"
password = "********"

username_input = WebDriverWait(driver, 10).until(
    EC.presence_of_element_located((By.NAME, "username"))
)
password_input = WebDriverWait(driver, 10).until(
    EC.presence_of_element_located((By.NAME, "password"))
)

username_input.send_keys(username)
password_input.send_keys(password)

password_input.send_keys(Keys.RETURN)

time.sleep(3)

for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
    instagram_link , test = row

    driver.get(instagram_link)

    excell_username = os.path.basename(urlparse(instagram_link).path)
    excell_username_selector = f'a[href="/{excell_username}/followers/"] span[class="_ac2a"] span'

    try:
        excell_username_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, excell_username_selector))
        )

        follower_count = excell_username_element.text.strip()
        followers[instagram_link] = follower_count
    except StaleElementReferenceException:
        print("Element has become stale. Retrying...")
        time.sleep(2)
        excell_username_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, excell_username_selector))
        )
        follower_count = excell_username_element.text.strip()
        followers[instagram_link] = follower_count
    except TimeoutException:
        print(f"Timeout exception. Element not found for link: {instagram_link}")
        followers[instagram_link] = None
    except Exception as e:
        print(f"Hata: {e}")
        followers[instagram_link] = None

driver.quit()

new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active

new_sheet['A1'] = 'Kullanıcı Adı'
new_sheet['B1'] = 'Instagram Link'
new_sheet['C1'] = 'Takipçi Sayısı'

row_index = 2
for instagram_link, takipci_sayisi in followers.items():
    kullanici_adi = os.path.basename(urlparse(instagram_link).path)

    new_sheet.cell(row=row_index, column=1, value=kullanici_adi)
    new_sheet.cell(row=row_index, column=2, value=instagram_link)
    new_sheet.cell(row=row_index, column=3, value=takipci_sayisi)
    row_index += 1

new_workbook.save('followers.xlsx')
